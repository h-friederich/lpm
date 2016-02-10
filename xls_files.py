# -*- coding: utf-8 -*-
"""
Excel-file importing module for lpm

A two-stage mechanism is used for importing files. In the first step the file is uploaded and stored in a temporary
location. The file is then parsed and presented for validation. When the user accepts the validated data the file
is re-read and imported to the database.
This module only provides the low-level functionality for storing and parsing file data.

Note: There is no protection against the temporary file being removed between upload and validation.

:copyright: (c) 2016 Hannes Friederich.
:license: BSD, see LICENSE for more details.
"""

import os
import tempfile
from werkzeug import secure_filename
from flask import request
from flask_wtf import Form
from flask_wtf.file import FileField
from wtforms import HiddenField
from openpyxl import load_workbook


class FileForm(Form):
    """
    Helper class for two-stage file imports
    In a first stage, the file is submitted and stored in a temporary location, parsed and shown to the user for review
    In a second stage, the data is imported into the database itself.
    The filename parameter stores the temporary file name
    """
    file = FileField(label='File')
    tmpname = HiddenField()


def read_xls(filename):
    """
    Attempts to read the given Excel-File and returns the data as a tuple (headers, list of dicts)
    The first row is considered the header row.
    Parsing stops at the first column with empty header.
    """
    wb = load_workbook(filename, read_only=True)
    if len(wb.sheetnames) > 1:
        raise ValueError('only single-sheet workbooks supported')
    ws = wb.active  # active sheet
    headers = None
    outdata = list()
    for row in ws.iter_rows():
        if not headers:
            headers = list()
            for cell in row:
                if not cell.value:
                    break
                headers.append(str(cell.value).strip())
        else:
            data = dict()
            for idx, cell in enumerate(row):
                # protection in case there are more data columns than header columns
                if idx >= len(headers):
                    break
                value = cell.value
                if isinstance(value, bool):
                    data[headers[idx]] = value
                elif value:
                    data[headers[idx]] = value
            if data:  # otherwise empty row
                outdata.append(data)
    return headers, outdata


def save_to_tmp(form):
    """
    Saves the file from the given file form into /tmp using a temporary name which preserves the suffix,
    stores the temporary name in the form, and reads the file content.
    WTF is NOT used for the file handling, since the file upload handling seems broken.
    """
    file = request.files.get('file')
    suffix = os.path.splitext(secure_filename(file.filename))[-1]
    tf = tempfile.NamedTemporaryFile(dir='/tmp', delete=False, suffix=suffix, prefix='lpm_tmp_')
    filepath = tf.name
    tf.close()
    file.save(filepath)
    form.tmpname.data = os.path.basename(filepath)
    return filepath


def extract_filepath(form):
    """
    Returns the file path represented by the form
    """
    return os.path.join('/tmp/', secure_filename(form.tmpname.data))